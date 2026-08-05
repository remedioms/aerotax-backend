#!/usr/bin/env python3
"""Edelweiss-Logbook-XLSX -> AeroX-Flugbuch-JSON.

    python3 parse_edw_xlsx.py <quelle.xlsx> <ziel.json>

Der EDW-Export unterscheidet bewusst zwischen der gesamten ``Flight Time``
und der dem Crewmitglied gutgeschriebenen ``Function Time PIC/Copilot``.
Gerade bei augmented Longhaul-Crews ist die Function Time nur 1/2 oder 2/3
der gesamten Blockzeit. Fuer das persoenliche Flugbuch darf daher nur die
Function Time importiert werden. Zeilen mit 00:00 in beiden Function-Spalten
sind belegte Beobachter-/Instruktor-Sektoren und werden nicht als eigene
Flugzeit erfunden.
"""

import argparse
import json
import os
import re
from datetime import date, datetime, time, timedelta, timezone

import openpyxl

from legkeys import dedupe_keys


EXPECTED_HEADERS = (
    "Date",
    "Departure Airport",
    "Departure TimeUtc",
    "Arrival Airport",
    "Arrival TimeUtc",
    "LH/SH",
    "Aircraft Type",
    "Aircraft Registration",
    "Takeoffs",
    "Landings",
    "Flight Time",
    "Function Time PIC",
    "Function Time Copilot",
    "Crew",
    "Crew Remarks",
)

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
AIRPORTS_PATH = os.path.join(ROOT, "airports_compact.json")


def airport_map(path=AIRPORTS_PATH):
    raw = json.load(open(path, encoding="utf-8"))
    fields = raw["fields"]
    iata_i, icao_i = fields.index("iata"), fields.index("icao")
    return {
        str(row[icao_i]).upper(): str(row[iata_i]).upper()
        for row in raw["rows"]
        if row[icao_i] and row[iata_i]
    }


ICAO_TO_IATA = airport_map()


def minutes(value):
    if value is None:
        raise ValueError("Zeitwert fehlt")
    if isinstance(value, timedelta):
        result = int(value.total_seconds() // 60)
    elif isinstance(value, time):
        result = value.hour * 60 + value.minute
    elif isinstance(value, (int, float)):
        result = round(value * 24 * 60)
    else:
        match = re.fullmatch(r"(\d{1,2}):(\d{2})", str(value).strip())
        if not match:
            raise ValueError(f"ungueltige Zeit: {value!r}")
        result = int(match.group(1)) * 60 + int(match.group(2))
    if not 0 <= result < 24 * 60:
        raise ValueError(f"Zeit ausserhalb eines Tages: {value!r}")
    return result


def source_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"ungueltiges Datum: {value!r}") from exc


def count(value, label):
    if isinstance(value, bool):
        raise ValueError(f"ungueltige {label}: {value!r}")
    try:
        parsed = int(value or 0)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"ungueltige {label}: {value!r}") from exc
    if parsed not in (0, 1):
        raise ValueError(f"{label} ausserhalb 0/1: {value!r}")
    return parsed


def airport(value):
    icao = str(value or "").strip().upper()
    if not re.fullmatch(r"[A-Z]{4}", icao):
        raise ValueError(f"ungueltiger ICAO-Flughafen: {value!r}")
    try:
        return ICAO_TO_IATA[icao]
    except KeyError as exc:
        raise ValueError(f"ICAO-Flughafen nicht aufloesbar: {icao}") from exc


def aircraft_type(value):
    raw = str(value or "").strip().upper()
    if not re.fullmatch(r"\d{3}", raw):
        raise ValueError(f"ungueltiger EDW-Flugzeugtyp: {value!r}")
    return "A" + raw


def parse_workbook(path):
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if book.sheetnames != ["Logbook"]:
        raise ValueError(f"unerwartete Tabellenblaetter: {book.sheetnames!r}")
    sheet = book["Logbook"]
    header = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
    if tuple(header[:len(EXPECTED_HEADERS)]) != EXPECTED_HEADERS:
        raise ValueError("EDW-Logbook-Spaltenkopf stimmt nicht")
    if any(value not in (None, "") for value in header[len(EXPECTED_HEADERS):]):
        raise ValueError("unerwartete Zusatzspalten im EDW-Logbook")

    legs = []
    skipped_zero = 0
    source_flight_min = source_credited_min = 0
    source_takeoffs = source_landings = 0

    for row_number, raw in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        row = tuple(raw[:len(EXPECTED_HEADERS)])
        if all(value in (None, "") for value in row):
            continue
        if len(row) != len(EXPECTED_HEADERS) or any(value is None for value in row[:13]):
            raise ValueError(f"unvollstaendige EDW-Zeile {row_number}")

        day = source_date(row[0])
        dep_min, arr_min = minutes(row[2]), minutes(row[4])
        flight_min = minutes(row[10])
        pic_min, copilot_min = minutes(row[11]), minutes(row[12])
        takeoffs = count(row[8], "Starts")
        landings = count(row[9], "Landungen")

        source_flight_min += flight_min
        source_takeoffs += takeoffs
        source_landings += landings

        if pic_min and copilot_min:
            raise ValueError(f"PIC und Copilot gleichzeitig belegt in Zeile {row_number}")
        credited = pic_min or copilot_min
        role = "PIC" if pic_min else "FO"

        if not credited:
            if flight_min or takeoffs or landings:
                raise ValueError(
                    f"Null-Funktionszeit mit Flugzeit/Start/Landung in Zeile {row_number}"
                )
            skipped_zero += 1
            continue
        if not 0 < credited <= flight_min:
            raise ValueError(f"unplausible Funktionszeit in Zeile {row_number}")

        clock_min = (arr_min - dep_min) % (24 * 60)
        if clock_min != flight_min:
            raise ValueError(
                f"Flight Time passt nicht zu UTC-Zeiten in Zeile {row_number}"
            )

        dep = datetime.combine(day, time(), tzinfo=timezone.utc) \
            + timedelta(minutes=dep_min)
        arr = datetime.combine(day, time(), tzinfo=timezone.utc) \
            + timedelta(minutes=arr_min)
        if arr <= dep:
            arr += timedelta(days=1)

        registration = str(row[7]).strip().upper()
        if not re.fullmatch(r"HB-[A-Z0-9]{3}", registration):
            raise ValueError(f"ungueltiges Kennzeichen in Zeile {row_number}")

        leg = {
            "date": day.isoformat(),
            "from": airport(row[1]),
            "to": airport(row[3]),
            "dep_iso": dep.strftime("%Y-%m-%dT%H:%M:00Z"),
            "arr_iso": arr.strftime("%Y-%m-%dT%H:%M:00Z"),
            "block_min": credited,
            "type": aircraft_type(row[6]),
            "reg": registration,
            "role": role,
        }
        # Der Export liefert nur Gesamtzaehler, keinen Tag/Nacht-Split.
        # Wie bei anderen belegten Legacy-Quellen bleiben die Summen erhalten;
        # ein Split wird nicht erfunden.
        if takeoffs:
            leg["to_day"] = takeoffs
        if landings:
            leg["ldg_day"] = landings
        legs.append(leg)
        source_credited_min += credited

    if not legs:
        raise ValueError("EDW-Logbook enthaelt keine anrechenbaren Flugzeilen")
    legs.sort(key=lambda leg: (leg["date"], leg["dep_iso"]))
    collisions = dedupe_keys(legs)

    if sum(leg["block_min"] for leg in legs) != source_credited_min:
        raise ValueError("Funktionszeit-Kontrollsumme weicht ab")
    if sum(leg.get("to_day", 0) for leg in legs) != source_takeoffs:
        raise ValueError("Start-Kontrollsumme weicht ab")
    if sum(leg.get("ldg_day", 0) for leg in legs) != source_landings:
        raise ValueError("Landungs-Kontrollsumme weicht ab")

    controls = {
        "source_rows": len(legs) + skipped_zero,
        "skipped_zero_function_rows": skipped_zero,
        "source_flight_min": source_flight_min,
        "credited_min": source_credited_min,
        "takeoffs": source_takeoffs,
        "landings": source_landings,
        "key_collisions": len(collisions),
    }
    return {"legs": legs, "sim": []}, controls


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source")
    parser.add_argument("destination")
    args = parser.parse_args()
    result, controls = parse_workbook(args.source)
    with open(args.destination, "w", encoding="utf-8") as handle:
        json.dump(result, handle, ensure_ascii=False)

    legs = result["legs"]
    total = controls["credited_min"]
    print(f"Legs      : {len(legs)}   Block {total // 60}:{total % 60:02d}")
    print(f"Landungen : {controls['landings']}   Starts {controls['takeoffs']}")
    print(f"Sims      : 0")
    print(f"Zeitraum  : {legs[0]['date']} -> {legs[-1]['date']}")
    print(f"Nullzeit  : {controls['skipped_zero_function_rows']} belegt uebersprungen")
    print(f"Key-Schutz: {controls['key_collisions']} kollidierende Legs nummeriert")


if __name__ == "__main__":
    main()
